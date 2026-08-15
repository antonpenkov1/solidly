#!/usr/bin/env python3
"""Regenerates Spoonlet/Core/WHOStandards.swift from the WHO Child Growth Standards.

Downloads the published z-score spreadsheets from who.int and writes the L, M and S
parameters for months 0-24 as Swift literals. Those tables carry LMS directly, so nothing
is fitted or interpolated here — GrowthMathTests reconstructs the published ±2 and ±3 SD
lines from what this writes, which is what catches a bad regeneration.

    pip install openpyxl
    python3 Tools/generate_who.py

The `sfvrsn` query parameters are cache-busters WHO rotates. If a download 404s, open
https://www.who.int/tools/child-growth-standards/standards and copy the current Excel link
for that indicator.
"""
import io
import sys
import urllib.request
from pathlib import Path

import openpyxl

BASE = ("https://cdn.who.int/media/docs/default-source/child-growth/"
        "child-growth-standards/indicators")

SOURCES = [
    ("weight", "boy",
     f"{BASE}/weight-for-age/wfa_boys_0-to-5-years_zscores.xlsx?sfvrsn=97a05331_9"),
    ("weight", "girl",
     f"{BASE}/weight-for-age/wfa_girls_0-to-5-years_zscores.xlsx?sfvrsn=4c03b8db_7"),
    ("length", "boy",
     f"{BASE}/length-height-for-age/lhfa_boys_0-to-2-years_zscores.xlsx?sfvrsn=30e044c_9"),
    ("length", "girl",
     f"{BASE}/length-height-for-age/lhfa_girls_0-to-2-years_zscores.xlsx?sfvrsn=e9e66a95_11"),
    ("head", "boy",
     f"{BASE}/head-circumference-for-age/hcfa-boys-0-5-zscores.xlsx?sfvrsn=701b4c5f_9"),
    ("head", "girl",
     f"{BASE}/head-circumference-for-age/hcfa-girls-0-5-zscores.xlsx?sfvrsn=8f959f88_6"),
]

MAX_MONTH = 24


def fetch(url: str) -> list[tuple[int, float, float, float]]:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = response.read()

    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    sheet = workbook.active

    header = [str(cell) if cell is not None else ""
              for cell in next(sheet.iter_rows(values_only=True))]
    for index, expected in enumerate(("Month", "L", "M", "S")):
        if header[index] != expected:
            raise SystemExit(f"unexpected column layout in {url}: {header}")

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        month = row[0]
        if month is None or int(month) > MAX_MONTH:
            continue
        rows.append((int(month), float(row[1]), float(row[2]), float(row[3])))

    if len(rows) != MAX_MONTH + 1:
        raise SystemExit(f"expected {MAX_MONTH + 1} rows from {url}, got {len(rows)}")
    return rows


def main() -> int:
    lines = [
        "// Generated from the WHO Child Growth Standards (2006) z-score tables.",
        "// Source files: https://www.who.int/tools/child-growth-standards/standards",
        "// Columns are the published Box-Cox parameters L (power), M (median) and S "
        "(coefficient of variation)",
        "// for months 0-24. Do not hand-edit - regenerate with Tools/generate_who.py.",
        "",
        "import Foundation",
        "",
        "/// One row of the WHO LMS reference: the skewness, median and variation of the",
        "/// measurement distribution at a whole month of age.",
        "struct LMS {",
        "    let month: Int",
        "    let l: Double",
        "    let m: Double",
        "    let s: Double",
        "}",
        "",
        "enum WHOStandards {",
    ]

    for indicator, sex, url in SOURCES:
        rows = fetch(url)
        name = f"{indicator}{sex.capitalize()}s"
        print(f"  {name}: {len(rows)} rows", file=sys.stderr)
        lines.append(f"    /// {indicator.capitalize()}-for-age, {sex}s, 0-24 months. "
                     "WHO Child Growth Standards 2006.")
        lines.append(f"    static let {name}: [LMS] = [")
        for month, l, m, s in rows:
            lines.append(f"        LMS(month: {month}, l: {l}, m: {m}, s: {s}),")
        lines.append("    ]")
        lines.append("")

    lines.append("""    static func table(for indicator: GrowthIndicator, sex: ChildSex) -> [LMS] {
        switch (indicator, sex) {
        case (.weight, .boy): return weightBoys
        case (.weight, .girl): return weightGirls
        case (.length, .boy): return lengthBoys
        case (.length, .girl): return lengthGirls
        case (.head, .boy): return headBoys
        case (.head, .girl): return headGirls
        }
    }
}""")

    out = Path(__file__).resolve().parent.parent / "Spoonlet" / "Core" / "WHOStandards.swift"
    out.write_text("\n".join(lines) + "\n")
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
